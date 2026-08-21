# -*- coding: utf-8 -*-
"""Merger classes for handling merged cells, images, data validation, and auto filter."""

from collections import defaultdict
from copy import copy

from openpyxl.worksheet.cell_range import CellRange, MultiCellRange

from .image import Img


class MergeMixin:
    """Base mixin for merge tracking logic."""

    def set_range(self, rdrowx=-1, rdcolx=-1, wtrowx=-1, wtcolx=-1):
        """Set the range positions for tracking."""
        self.start_rdrowx = rdrowx
        self.start_rdcolx = rdcolx
        self.start_wtrowx = wtrowx
        self.start_wtcolx = wtcolx
        self.end_wtrowx = wtrowx
        self.end_wtcolx = wtcolx

    def is_in_range(self, rdrowx, rdcolx):
        """Check if cell is in the merged range."""
        return (
            self._first_row <= rdrowx <= self._last_row
            and self._first_col <= rdcolx <= self._last_col
        )

    def to_be_merged(self, rdrowx, rdcolx):
        """Check if cell should be merged with current range."""
        if rdrowx > self.start_rdrowx:
            return True
        else:
            return rdrowx == self.start_rdrowx and rdcolx > self.start_rdcolx

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        """Track cell position for merging."""
        if not self.is_in_range(rdrowx, rdcolx):
            return False
        if self.start_rdrowx == -1:
            self.set_range(rdrowx, rdcolx, wtrowx, wtcolx)
        elif self.to_be_merged(rdrowx, rdcolx):
            self.end_wtrowx = max(self.end_wtrowx, wtrowx)
            self.end_wtcolx = max(self.end_wtcolx, wtcolx)
        else:
            self.new_range()
            self.set_range(rdrowx, rdcolx, wtrowx, wtcolx)
        return True

    def new_range(self):
        """Create new range (implemented by subclasses)."""
        pass

    def collect_range(self):
        """Finalize range collection."""
        self.new_range()
        self.set_range()


class MergerMixin:
    """Mixin for merger classes."""

    @property
    def to_merge(self):
        """Check if there are items to merge."""
        return bool(self._merge_list)


class CellMerge(MergeMixin):
    def __init__(self, cell_range, merger):
        self.merger = merger
        self.set_range()
        self._first_row = cell_range.min_row
        self._last_row = cell_range.max_row
        self._first_col = cell_range.min_col
        self._last_col = cell_range.max_col

    def new_range(self):
        if (
            self.start_wtrowx == self.end_wtrowx
            and self.start_wtcolx == self.end_wtcolx
        ):
            return
        range = CellRange(
            None, self.start_wtcolx, self.start_wtrowx, self.end_wtcolx, self.end_wtrowx
        )
        self.merger.add_new_range(range)


class CellMerger(MergerMixin):
    def __init__(self, sheet):
        self.range_list = []
        self._merge_list = []
        self.get_merge_list(sheet)

    def get_merge_list(self, sheet):
        for range in sheet.merged_cells:
            _merge = CellMerge(range, self)
            self._merge_list.append(_merge)

    def add_new_range(self, range):
        self.range_list.append(range)

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        for _merge in self._merge_list:
            is_in_range = _merge.merge_cell(rdrowx, rdcolx, wtrowx, wtcolx)
            if is_in_range:
                break

    def collect_range(self, wtsheet):
        for _merge in self._merge_list:
            _merge.collect_range()
        for range in self.range_list:
            wtsheet.merged_cells.add(range)
        self.range_list.clear()


class DataValidation(MergeMixin):
    def __init__(self, cell_range, merger, dv_key):
        self.dv_key = dv_key
        self.merger = merger
        self.set_range()
        self._first_row = cell_range.min_row
        self._last_row = cell_range.max_row
        self._first_col = cell_range.min_col
        self._last_col = cell_range.max_col

    def new_range(self):
        if self.start_wtrowx == -1:
            return
        range = CellRange(
            None, self.start_wtcolx, self.start_wtrowx, self.end_wtcolx, self.end_wtrowx
        )
        self.merger.add_new_range(self.dv_key, range)


class DvMerger(MergerMixin):
    def __init__(self, sheet):
        self.dv_map = {}
        self.dv_copy_map = {}
        self._merge_list = []
        self.get_merge_list(sheet)

    def get_merge_list(self, rdsheet):
        for index, dv in enumerate(rdsheet.data_validations.dataValidation):
            self.dv_map[index] = dv
            for crange in dv.ranges:
                _merge = DataValidation(crange, self, index)
                self._merge_list.append(_merge)

    def add_new_range(self, dv_key, range):
        dv_copy = self.dv_copy_map.get(dv_key)
        if not dv_copy:
            dv_copy = copy(self.dv_map[dv_key])
            dv_copy.ranges = MultiCellRange()
            self.dv_copy_map[dv_key] = dv_copy
        dv_copy.ranges.add(range)

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        for _merge in self._merge_list:
            is_in_range = _merge.merge_cell(rdrowx, rdcolx, wtrowx, wtcolx)
            if is_in_range:
                break

    def collect_range(self, wtsheet):
        for _merge in self._merge_list:
            _merge.collect_range()
        for key, dv in self.dv_copy_map.items():
            wtsheet.data_validations.append(dv)
        self.dv_copy_map.clear()


from collections import defaultdict


class ImageMerge(MergeMixin):
    def __init__(self, image, merger, image_count_dict):
        self.merger = merger
        self.image = image
        self.set_range()
        self.image_copy_map = {}
        self.image_ref_map = {}
        _from = image.anchor._from
        _to = image.anchor.to
        self._first_row = rlo = _from.row + 1
        self._first_col = clo = _from.col + 1
        self._last_row = rhi = _to.row + 1
        self._last_col = chi = _to.col + 1
        _top_left = (rlo, clo)
        count = image_count_dict[_top_left]
        image_count_dict[_top_left] += 1
        self.image_key = (rlo, clo, count)

    def new_range(self):
        if self.start_wtrowx == -1:
            return
        image = Img(self.image)
        _from = image.anchor._from
        _to = image.anchor.to
        _from.row = self.start_wtrowx - 1
        _from.col = self.start_wtcolx - 1
        _to.row = self.end_wtrowx - 1
        _to.col = self.end_wtcolx - 1
        self.image_copy_map[(self.start_wtrowx, self.start_wtcolx)] = image

    def set_image_ref(self, image_ref):
        if image_ref.image:
            self.image_ref_map[image_ref.wt_top_left] = image_ref.image

    def collect_range(self):
        self.new_range()
        self.set_range()
        for key, image in self.image_copy_map.items():
            ref = self.image_ref_map.get(key)
            if ref:
                image.set_ref(ref)
            self.merger.add_image(image)
        self.image_copy_map.clear()
        self.image_ref_map.clear()


class ImageMerger(MergerMixin):
    def __init__(self, sheet):
        self.images = []
        self._merge_map = {}
        self._merge_list = []
        self.max_row = 0
        self.max_col = 0
        self.get_merge_list(sheet)

    def get_merge_list(self, rdsheet):
        image_count_dict = defaultdict(int)
        for image in rdsheet._images:
            # print(image.ref, id(image.ref))
            _merge = ImageMerge(image, self, image_count_dict)
            self._merge_map[_merge.image_key] = _merge
            self._merge_list.append(_merge)
            self.max_row = max(self.max_row, _merge._last_row)
            self.max_col = max(self.max_col, _merge._last_col)

    def add_image(self, image):
        self.images.append(image)

    def set_image_ref(self, image_ref):
        _merge = self._merge_map.get(image_ref.image_key)
        if not _merge:
            return False
        _merge.set_image_ref(image_ref)
        return True

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        for _merge in self._merge_list:
            _merge.merge_cell(rdrowx, rdcolx, wtrowx, wtcolx)

    def collect_range(self, wtsheet):
        for _merge in self._merge_list:
            _merge.collect_range()
        wtsheet._images = self.images
        self.images = []


class AutoFilter(MergeMixin):
    def __init__(self, rdsheet):
        if not rdsheet.auto_filter.ref:
            self.to_merge = False
            return
        self.to_merge = True
        self.auto_filter = rdsheet.auto_filter
        self.set_range()
        cell_range = CellRange(rdsheet.auto_filter.ref)
        self._first_row = cell_range.min_row
        self._last_row = cell_range.max_row
        self._first_col = cell_range.min_col
        self._last_col = cell_range.max_col
        self.first_af = None

    def new_range(self):
        if self.start_wtrowx == -1:
            return
        if not self.first_af:
            self.first_af = CellRange(
                None,
                self.start_wtcolx,
                self.start_wtrowx,
                self.end_wtcolx,
                self.end_wtrowx,
            )

    def collect_range(self, wtsheet):
        self.new_range()
        self.set_range()
        if wtsheet.auto_filter.ref:
            self.first_af = None
            return
        if self.first_af:
            wtsheet.auto_filter = copy(self.auto_filter)
            wtsheet.auto_filter.ref = self.first_af.coord
            self.first_af = None


class DefinedName(MergeMixin):
    pass


class DefinedNames(MergerMixin):
    pass


class Merger:
    def __init__(self, rdsheet):
        cell_merger = CellMerger(rdsheet)
        dv_merger = DvMerger(rdsheet)
        self.image_merger = image_merger = ImageMerger(rdsheet)
        auto_filter = AutoFilter(rdsheet)
        _merger_list = [cell_merger, dv_merger, image_merger, auto_filter]
        self.merger_list = []
        for merger in _merger_list:
            if merger.to_merge:
                self.merger_list.append(merger)
        self._extra_images = []

    def merge_cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        for merger in self.merger_list:
            merger.merge_cell(rdrowx, rdcolx, wtrowx, wtcolx)

    def collect_range(self, wtsheet):
        for merger in self.merger_list:
            merger.collect_range(wtsheet)
        self._flush_extra_images(wtsheet)

    def _flush_extra_images(self, wtsheet):
        from math import floor
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.units import points_to_pixels, pixels_to_EMU

        _DEFAULT_COL_WIDTH = 8.43
        _DEFAULT_ROW_HEIGHT = 15

        for image_ref in self._extra_images:
            img = image_ref.image
            if not img:
                continue
            row = image_ref.wtrowx or 1
            col = image_ref.wtcolx or 1
            col_letter = get_column_letter(col)
            col_dim = wtsheet.column_dimensions.get(col_letter)
            row_dim = wtsheet.row_dimensions.get(row)
            col_width = (col_dim.width if col_dim else None) or _DEFAULT_COL_WIDTH
            row_height = (row_dim.height if row_dim else None) or _DEFAULT_ROW_HEIGHT
            w = max(int(floor((col_width * 256 + 128) / 256 * 7)), 1)
            h = max(int(points_to_pixels(row_height)), 1)
            opx_img = OpenpyxlImage(img)
            opx_img.width = w
            opx_img.height = h
            opx_img.anchor = OneCellAnchor(
                _from=AnchorMarker(col=col, colOff=0, row=row - 1, rowOff=0),
                ext=XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h)),
            )
            wtsheet.add_image(opx_img)
        self._extra_images.clear()

    def set_image_ref(self, image_ref):
        handled = self.image_merger.set_image_ref(image_ref)
        if handled:
            return
        if getattr(image_ref, 'allow_insert', False) and image_ref.image:
            self._extra_images.append(image_ref)
