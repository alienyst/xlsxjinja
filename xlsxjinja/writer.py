# -*- coding: utf-8 -*-
"""
Main entry point for Excel template rendering.

This module provides the BookWriter class, which is the primary interface
for rendering Excel templates using Jinja2 syntax.

Flow:
1. Load Excel template file
2. Parse template into node tree structure
3. Render template with user data
4. Generate output Excel file
"""

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

from .base import BookBase, SheetBase
from .celltag import CellTag
from .config import config
from .image import img_cache
from .jinja import JinjaEnv
from .logger import get_logger, setup_logger
from .merger import Merger
from .nodemap import NodeMap
from .patch import *
from .richtexthandler import rich_handler
from .sheetresource import SheetResourceMap
from .utils import parse_cell_tag, tag_test
from .writermixin import BookMixin, Box, SheetMixin
from .xlnode import Cell, EmptyCell, Node, Row, Tree, create_cell


class SheetWriter(SheetBase, SheetMixin):
    """
    Handles writing of a single worksheet.

    Flow:
    1. Initialize with template sheet and target workbook
    2. Create new output sheet
    3. Copy sheet settings from template
    4. Track written rows/columns to avoid duplication
    5. Process cells and write to output sheet
    """

    def __init__(self, bookwriter, sheet_resource, sheet_name):
        """
        Initialize sheet writer.

        Args:
            bookwriter: Parent BookWriter instance
            sheet_resource: SheetResource with template data
            sheet_name: Name for output sheet

        Flow:
        1. Store references to workbook and template sheet
        2. Create new output sheet with given name
        3. Copy all sheet settings (format, print, margins, etc.)
        4. Initialize tracking sets for rows/columns
        5. Initialize box for region tracking
        """
        self.workbook = bookwriter.workbook
        self.merger = sheet_resource.merger
        self.rdsheet = sheet_resource.rdsheet
        self.wtsheet = self.workbook.create_sheet(title=sheet_name)
        self.copy_sheet_settings()
        self.wtrows = set()
        self.wtcols = set()
        self.box = Box(0, 0)


class BookWriter(BookBase, BookMixin):
    """
    Main interface for Excel template rendering with Jinja2.

    Usage:
        writer = BookWriter('template.xlsx', debug=True)
        writer.render_book(payloads)
        writer.save('output.xlsx')

    Flow:
    1. Load template Excel file
    2. Parse sheets into node trees
    3. Compile Jinja2 templates with node references
    4. Render with user data
    5. Write output Excel file
    """

    sheet_writer_cls = SheetWriter

    def __init__(self, fname, debug=False):
        """
        Initialize BookWriter and load template.

        Args:
            fname: Path to template Excel file (.xlsx)
            debug: Enable debug logging (default: False)

        Flow:
        1. Setup logger with debug mode
        2. Store debug flag in config
        3. Load template file
        """
        # Setup logger based on debug mode
        setup_logger("xlsxjinja", debug=debug)
        self.logger = get_logger("xlsxjinja")

        config.debug = debug
        self.load(fname)

    def load(self, fname):
        """
        Load Excel template file and initialize internal structures.

        Args:
            fname: Path to template Excel file

        Flow:
        1. Load workbook with openpyxl (rich_text=True for formatting)
        2. Initialize font map for font caching
        3. Create NodeMap for tracking template nodes
        4. Create JinjaEnv with custom extensions
        5. Set merger class for handling merged cells
        6. Initialize sheet writer and resource maps
        7. Process all worksheets into resources
        8. Remove original template sheets from workbook
        """
        self.workbook = load_workbook(fname, rich_text=True)
        self.font_map = {}
        self.node_map = NodeMap()
        self.jinja_env = JinjaEnv(self.node_map)
        self.merger_cls = Merger
        self.sheet_writer_map = {}
        self.sheet_resource_map = SheetResourceMap(self, self.jinja_env)
        for index, rdsheet in enumerate(self.workbook.worksheets):
            self.sheet_resource_map.add(rdsheet, rdsheet.title, index)
            self.workbook.remove(rdsheet)

    def build(self, sheet, index, merger):
        """
        Build node tree from Excel template sheet.

        Args:
            sheet: Worksheet to build from
            index: Sheet index number
            merger: Merger instance for handling merged cells/images

        Returns:
            Tree: Root node of the built tree structure

        Flow:
        1. Create root Tree node for the sheet
        2. Calculate max rows/cols including images
        3. Iterate through all rows
        4. For each row, create Row node
        5. Iterate through all columns in row
        6. For each cell:
           a. Check if cell exists, create EmptyCell if not
           b. Parse cell comments for special tags
           c. Handle string cells (check for rich text)
           d. Detect Jinja2 templates in cell values
           e. Create appropriate cell node (Cell, TagCell, RichTagCell, XvCell)
           f. Attach cell tags if present
        7. Add terminal node to tree
        8. Return complete tree
        """
        tree = Tree(index, self.node_map)
        max_row = max(sheet.max_row, merger.image_merger.max_row)
        max_col = max(sheet.max_column, merger.image_merger.max_col)
        for rowx in range(1, max_row + 1):
            row_node = Row(rowx)
            tree.add_child(row_node)
            for colx in range(1, max_col + 1):
                sheet_cell = sheet._cells.get((rowx, colx))
                if not sheet_cell:
                    cell_node = EmptyCell(rowx, colx)
                    tree.add_child(cell_node)
                    continue
                cell_tag_map = None
                if sheet_cell.comment:
                    comment = sheet_cell.comment.text
                    if tag_test(comment):
                        _, cell_tag_map = parse_cell_tag(comment)
                value = sheet_cell._value
                data_type = sheet_cell.data_type
                if data_type == "s":
                    rich_text = None
                    if isinstance(value, CellRichText):
                        rich_text = value
                        value = str(rich_text)
                    if not tag_test(value):
                        if rich_text:
                            cell_node = Cell(
                                sheet_cell, rowx, colx, rich_text, data_type
                            )
                        else:
                            cell_node = Cell(sheet_cell, rowx, colx, value, data_type)
                    else:
                        font = self.get_font(sheet_cell._style.fontId)
                        cell_node = create_cell(
                            sheet_cell,
                            rowx,
                            colx,
                            value,
                            rich_text,
                            data_type,
                            font,
                            rich_handler,
                        )
                else:
                    cell_node = Cell(sheet_cell, rowx, colx, value, data_type)
                if cell_tag_map:
                    cell_tag = CellTag(cell_tag_map)
                    cell_node.extend_cell_tag(cell_tag)
                    if colx == 1:
                        row_node.cell_tag = cell_tag
                tree.add_child(cell_node)
        tree.add_child(Node())  # Terminal node
        return tree

    def cleanup_defined_names(self):
        """
        Remove invalid defined names from workbook.

        Flow:
        1. Clear custom document properties (they can cause errors)
        2. Get current sheet count
        3. Iterate through all defined names
        4. Keep only names with valid sheet references
        5. Update workbook with cleaned names

        Note: This prevents "invalid file" errors in Excel
        """
        self.workbook.custom_doc_props = ()
        # Custom Document Properties cause invalid file error
        sheet_cnt = len(self.workbook.worksheets)
        valid_names = {}
        for k, v in self.workbook.defined_names.items():
            if v.localSheetId:
                if int(v.localSheetId) < sheet_cnt:
                    valid_names[k] = v
            else:
                valid_names[k] = v
        self.workbook.defined_names = valid_names

    def save(self, fname):
        """
        Save rendered workbook to Excel file.

        Args:
            fname: Output file path

        Flow:
        1. Ensure workbook has an active sheet
        2. Clean up invalid defined names
        3. Save workbook to file
        4. Clear image cache to free memory
        5. Remove all sheets from workbook
        6. Clear sheet writer map

        Note: Cleanup steps free memory and prepare for next render
        """
        if not self.workbook.active:
            self.workbook.active = 0
        self.cleanup_defined_names()
        self.workbook.save(fname)
        img_cache.clear()
        for sheet in self.workbook.worksheets:
            self.workbook.remove(sheet)
        self.sheet_writer_map.clear()
