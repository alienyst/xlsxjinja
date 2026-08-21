# -*- coding: utf-8 -*-
"""
Base classes for sheet and book operations.

This module provides base functionality for reading template sheets
and writing output sheets, including copying cell properties,
row/column dimensions, and sheet settings.
"""

import copy

from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter

from .cellcontext import CellContext


class SheetBase:
    """
    Base class for sheet operations.

    Provides core functionality for:
    - Copying sheet settings (format, print options, margins, etc.)
    - Copying row and column dimensions
    - Writing cells with formatting preservation

    Flow:
    1. Copy settings from template sheet to output sheet
    2. Track which rows/columns have been written
    3. Copy dimensions only once per row/column
    4. Write cells with full formatting preservation
    """

    def copy_sheet_settings(self):
        """
        Copy all sheet settings from template to output sheet.

        Flow:
        1. Copy sheet format and properties
        2. Copy print settings (page setup, options, areas)
        3. Copy page margins and protection
        4. Copy header/footer settings
        5. Copy sheet views (zoom, panes, etc.)
        6. Copy images collection
        """
        self.wtsheet.sheet_format = copy.copy(self.rdsheet.sheet_format)
        self.wtsheet.sheet_properties = copy.copy(self.rdsheet.sheet_properties)
        # copy print settings
        self.wtsheet.page_setup = copy.copy(self.rdsheet.page_setup)
        self.wtsheet.print_options = copy.copy(self.rdsheet.print_options)
        self.wtsheet._print_rows = copy.copy(self.rdsheet._print_rows)
        self.wtsheet._print_cols = copy.copy(self.rdsheet._print_cols)
        self.wtsheet._print_area = copy.copy(self.rdsheet._print_area)
        self.wtsheet.page_margins = copy.copy(self.rdsheet.page_margins)
        self.wtsheet.protection = copy.copy(self.rdsheet.protection)
        self.wtsheet.HeaderFooter = copy.copy(self.rdsheet.HeaderFooter)
        self.wtsheet.views = copy.copy(self.rdsheet.views)
        self.wtsheet._images = copy.copy(self.rdsheet._images)

    def copy_row_dimension(self, rdrowx, wtrowx):
        """
        Copy row dimension (height, style) from template to output.

        Args:
            rdrowx: Template row index
            wtrowx: Output row index

        Flow:
        1. Check if row already copied (avoid duplicates)
        2. Get row dimension from template
        3. Copy dimension to output sheet
        4. Update worksheet reference
        5. Mark row as copied
        """
        if wtrowx in self.wtrows:
            return
        dim = self.rdsheet.row_dimensions.get(rdrowx)
        if dim:
            self.wtsheet.row_dimensions[wtrowx] = copy.copy(dim)
            self.wtsheet.row_dimensions[wtrowx].worksheet = self.wtsheet
            self.wtrows.add(wtrowx)

    def copy_col_dimension(self, rdcolx, wtcolx):
        """
        Copy column dimension (width, style) from template to output.

        Args:
            rdcolx: Template column index
            wtcolx: Output column index

        Flow:
        1. Check if column already copied (avoid duplicates)
        2. Get column letter and dimension from template
        3. Copy dimension to output sheet
        4. If column index changed, adjust min/max bounds
        5. Update worksheet reference
        6. Mark column as copied
        """
        if wtcolx in self.wtcols:
            return
        rdkey = get_column_letter(rdcolx)
        rddim = self.rdsheet.column_dimensions.get(rdkey)
        if not rddim:
            return
        wtdim = copy.copy(rddim)
        if rdcolx != wtcolx:
            wtkey = get_column_letter(wtcolx)
            wtdim.index = wtkey
            d = wtcolx - rdcolx
            wtdim.min += d
            wtdim.max += d
        else:
            wtkey = rdkey
        self.wtsheet.column_dimensions[wtkey] = wtdim
        self.wtsheet.column_dimensions[wtkey].worksheet = self.wtsheet
        self.wtcols.add(wtcolx)

    def _cell(
        self, source_cell, rdrowx, rdcolx, wtrowx, wtcolx, value=None, data_type=None
    ):
        """
        Internal method to write cell with formatting.

        Args:
            source_cell: Template cell to copy from
            rdrowx: Template row index
            rdcolx: Template column index
            wtrowx: Output row index
            wtcolx: Output column index
            value: New value (optional, uses source if None)
            data_type: Data type (optional)

        Returns:
            Target cell object

        Flow:
        1. Get or create target cell at output position
        2. Set cell value (from source or provided)
        3. Handle formulas (strings starting with =)
        4. Copy cell style if present
        5. Copy hyperlink if present
        6. Return target cell
        """
        target_cell = self.wtsheet.cell(column=wtcolx, row=wtrowx)
        if value is None:
            target_cell.value = source_cell._value
            target_cell.data_type = source_cell.data_type
        elif isinstance(value, str) and value.startswith("="):
            target_cell.value = value
        elif data_type:
            target_cell._value = value
            target_cell.data_type = data_type
        else:
            target_cell.value = value
        if source_cell.has_style:
            target_cell._style = copy.copy(source_cell._style)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy.copy(source_cell.hyperlink)
        return target_cell

    def cell(
        self, source_cell, rdrowx, rdcolx, wtrowx, wtcolx, value=None, data_type=None
    ):
        """
        Write cell with dimension copying.

        Args:
            source_cell: Template cell to copy from
            rdrowx: Template row index
            rdcolx: Template column index
            wtrowx: Output row index
            wtcolx: Output column index
            value: New value (optional)
            data_type: Data type (optional)

        Returns:
            Target cell object

        Flow:
        1. Copy row dimension from template
        2. Copy column dimension from template
        3. Write cell using _cell method
        4. Return target cell
        """
        self.copy_row_dimension(rdrowx, wtrowx)
        self.copy_col_dimension(rdcolx, wtcolx)
        return self._cell(source_cell, rdrowx, rdcolx, wtrowx, wtcolx, value, data_type)

    def get_cell_context(self, cell_node, rv, cty):
        """
        Create cell context for writing.

        Args:
            cell_node: Node representing the cell
            rv: Rendered value
            cty: Cell type

        Returns:
            CellContext instance

        Flow:
        1. Create CellContext with sheet writer and cell data
        2. CellContext handles lazy cell creation
        3. Return context for further processing
        """
        return CellContext(self, cell_node, rv, cty)


class BookBase:
    """
    Base class for workbook operations.

    Provides font management with caching to optimize
    InlineFont creation for rich text handling.
    """

    def get_font(self, fontId):
        """
        Get or create InlineFont from font ID with caching.

        Args:
            fontId: Font identifier from workbook font table

        Returns:
            InlineFont instance

        Flow:
        1. Check if font already cached
        2. If cached, return existing InlineFont
        3. If not cached:
           a. Get font from workbook font table
           b. Create new InlineFont instance
           c. Copy all font properties (name, size, bold, italic, etc.)
           d. Cache for future use
           e. Return InlineFont

        Note: Caching improves performance for repeated font access
        """
        ifont = self.font_map.get(fontId)
        if ifont:
            return ifont
        else:
            font = self.workbook._fonts[fontId]
            ifont = InlineFont()
            ifont.rFont = font.name
            ifont.charset = font.charset
            ifont.family = font.family
            ifont.b = font.b
            ifont.i = font.i
            ifont.strike = font.strike
            ifont.outline = font.outline
            ifont.shadow = font.shadow
            ifont.condense = font.condense
            ifont.extend = font.extend
            ifont.color = font.color
            ifont.sz = font.sz
            ifont.u = font.u
            ifont.vertAlign = font.vertAlign
            ifont.scheme = font.scheme
            self.font_map[fontId] = ifont
            return ifont
